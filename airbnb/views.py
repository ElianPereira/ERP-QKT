"""
Vistas del módulo Airbnb
========================
Vistas para calendario unificado, reportes, iCal inverso y bloqueo manual.
"""
import calendar
import logging
from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import permission_required
from django.db.models import Count, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date

from core_erp.ratelimit import rate_limit

from .models import AnuncioAirbnb, ConflictoCalendario, PagoAirbnb, ReservaAirbnb

logger = logging.getLogger(__name__)

# Nombres de los meses, compartidos por el reporte fiscal y la conciliación de
# depósitos.
MESES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
}


def _construir_eventos_calendario(fecha_inicio, fecha_fin):
    """Arma la lista de eventos del calendario unificado, acotada al rango
    [fecha_inicio, fecha_fin) — nunca el histórico completo. La vista
    original consultaba todas las cotizaciones/reservas/asignaciones que
    hayan existido jamás en cada carga de página (SEC-DOS-001): con el
    histórico creciendo sin límite, eso degrada progresivamente."""
    from comercial.models import AsignacionEspacio, AsignacionPersonal, Cotizacion

    eventos_lista = []

    # EVENTOS DE LA QUINTA (Cotizaciones)
    cotizaciones = Cotizacion.objects.exclude(estado='CANCELADA').filter(
        fecha_evento__gte=fecha_inicio, fecha_evento__lt=fecha_fin,
    ).select_related('cliente')

    for c in cotizaciones:
        if c.estado == 'CONFIRMADA':
            color = '#27ae60'
            icon = ''
        else:
            color = '#95a5a6'
            icon = ''

        eventos_lista.append({
            'title': f"{icon} {c.cliente.nombre} - {c.nombre_evento}",
            'start': c.fecha_evento.strftime("%Y-%m-%d"),
            'color': color,
            'url': f'/admin/comercial/cotizacion/{c.id}/change/',
            'extendedProps': {'tipo': 'evento'}
        })

    # RESERVAS DE AIRBNB — se incluye cualquier reserva cuyo rango se
    # traslape con [fecha_inicio, fecha_fin), no solo las que empiezan ahí.
    reservas = ReservaAirbnb.objects.filter(
        estado__in=['CONFIRMADA', 'BLOQUEADA', 'PENDIENTE'],
        anuncio__activo=True,
        fecha_inicio__lt=fecha_fin, fecha_fin__gte=fecha_inicio,
    ).select_related('anuncio')

    for r in reservas:
        tiene_conflicto = r.conflictos.filter(estado='PENDIENTE').exists()

        if tiene_conflicto:
            color = '#e74c3c'
            icon = ''
        elif r.estado == 'PENDIENTE':
            color = '#f39c12'
            icon = ''
        elif r.estado == 'BLOQUEADA':
            color = '#6c757d'
            icon = ''
        elif r.anuncio.tipo == 'CASA':
            color = '#3498db'
            icon = ''
        else:
            color = '#e67e22'
            icon = ''

        eventos_lista.append({
            'title': f"{icon} {r.anuncio.nombre}: {r.titulo or 'Reserva'}",
            'start': r.fecha_inicio.strftime("%Y-%m-%d"),
            'end': r.fecha_fin.strftime("%Y-%m-%d"),
            'color': color,
            'url': f'/admin/airbnb/reservaairbnb/{r.id}/change/',
            'extendedProps': {'tipo': 'airbnb'}
        })

    # ASIGNACIONES DE ESPACIO
    asignaciones_esp = AsignacionEspacio.objects.filter(
        fecha__gte=fecha_inicio, fecha__lt=fecha_fin,
    ).select_related('espacio', 'cotizacion__cliente')
    for a in asignaciones_esp:
        eventos_lista.append({
            'title': f"📍 {a.espacio.nombre}: COT-{a.cotizacion_id:03d}",
            'start': f"{a.fecha.strftime('%Y-%m-%d')}T{a.hora_inicio.strftime('%H:%M:%S')}",
            'end': f"{a.fecha.strftime('%Y-%m-%d')}T{a.hora_fin.strftime('%H:%M:%S')}" if a.hora_fin > a.hora_inicio else None,
            'color': '#9b59b6',
            'url': f'/admin/comercial/asignacionespacio/{a.id}/change/',
            'extendedProps': {'tipo': 'espacio'}
        })

    # ASIGNACIONES DE PERSONAL
    asignaciones_per = AsignacionPersonal.objects.filter(
        fecha__gte=fecha_inicio, fecha__lt=fecha_fin,
    ).select_related('empleado')
    for a in asignaciones_per:
        eventos_lista.append({
            'title': f"👤 {a.empleado.nombre} ({a.get_rol_display()})",
            'start': f"{a.fecha.strftime('%Y-%m-%d')}T{a.hora_inicio.strftime('%H:%M:%S')}",
            'color': '#16a085',
            'url': f'/admin/comercial/asignacionpersonal/{a.id}/change/',
            'extendedProps': {'tipo': 'personal'}
        })

    return eventos_lista


@staff_member_required
@permission_required('airbnb.view_reservaairbnb', raise_exception=True)
def calendario_unificado(request):
    """
    Calendario unificado que muestra eventos de la quinta + reservas de Airbnb.

    La página en sí no trae eventos: FullCalendar los pide por AJAX a
    `calendario_unificado_eventos` para el rango que esté visible en cada
    momento (mes/semana actual, o el que el usuario navegue a continuación),
    en vez de recibir el histórico completo cada vez que se abre la página.
    """
    conflictos_pendientes = ConflictoCalendario.objects.filter(estado='PENDIENTE').count()

    # URL de iCal para mostrar en la página
    ical_url = request.build_absolute_uri('/airbnb/ical/eventos/')

    context = {
        'eventos_url': reverse('calendario_unificado_eventos'),
        'conflictos_pendientes': conflictos_pendientes,
        'ical_url': ical_url,
        'title': 'Calendario Unificado',
    }

    return render(request, 'admin/airbnb/calendario_unificado.html', context)


@staff_member_required
@permission_required('airbnb.view_reservaairbnb', raise_exception=True)
def calendario_unificado_eventos(request):
    """JSON de eventos del calendario, acotado a `start`/`end` (YYYY-MM-DD,
    fin exclusivo) — el rango que FullCalendar tenga visible en cada momento.
    Devuelve un array plano de eventos, no un objeto envolvente: es el
    formato que FullCalendar espera cuando `events` recibe una función que
    llama a `successCallback(eventos)`."""
    fecha_inicio = parse_date(request.GET.get('start', ''))
    fecha_fin = parse_date(request.GET.get('end', ''))
    if not fecha_inicio or not fecha_fin:
        return JsonResponse({'error': "Parámetros 'start' y 'end' requeridos (YYYY-MM-DD)."}, status=400)

    return JsonResponse(_construir_eventos_calendario(fecha_inicio, fecha_fin), safe=False)


@staff_member_required
@permission_required('airbnb.view_pagoairbnb', raise_exception=True)
def reporte_pagos_airbnb(request):
    """
    Reporte de pagos de Airbnb para el contador.
    """
    hoy = timezone.localdate()

    try:
        año = int(str(request.GET.get('año', hoy.year)).replace(',', '').strip())
    except (TypeError, ValueError):
        año = hoy.year

    pagos = PagoAirbnb.objects.filter(estado='PAGADO')

    if año:
        pagos = pagos.filter(fecha_pago__year=año)

    mes = request.GET.get('mes', '')
    if mes:
        try:
            numero_mes = int(mes)
        except (TypeError, ValueError):
            numero_mes = None
        if numero_mes and 1 <= numero_mes <= 12:
            pagos = pagos.filter(fecha_pago__month=numero_mes)

    totales = pagos.aggregate(
        total_bruto=Sum('monto_bruto'),
        total_comision=Sum('comision_airbnb'),
        total_isr=Sum('retencion_isr'),
        total_iva=Sum('retencion_iva'),
        total_neto=Sum('monto_neto'),
        num_reservas=Count('id'),
    )

    for key in totales:
        if totales[key] is None:
            totales[key] = Decimal('0.00') if 'total' in key else 0

    resumen_mensual = pagos.annotate(
        mes=TruncMonth('fecha_checkin')
    ).values('mes').annotate(
        bruto=Sum('monto_bruto'),
        neto=Sum('monto_neto'),
        isr=Sum('retencion_isr'),
        iva=Sum('retencion_iva'),
        reservas=Count('id'),
    ).order_by('mes')

    resumen_anuncio = pagos.values(
        'anuncio__nombre'
    ).annotate(
        bruto=Sum('monto_bruto'),
        neto=Sum('monto_neto'),
        reservas=Count('id'),
    ).order_by('-bruto')

    context = {
        'pagos': pagos.select_related('anuncio').order_by('-fecha_checkin'),
        'totales': totales,
        'resumen_mensual': resumen_mensual,
        'resumen_anuncio': resumen_anuncio,
        'año_actual': año,
        'mes_actual': int(mes) if mes else None,
        'años_disponibles': range(2024, hoy.year + 2),
        'title': 'Reporte de Pagos Airbnb',
    }

    if request.GET.get('export') == 'excel':
        return exportar_reporte_excel(pagos, totales, año)

    return render(request, 'admin/airbnb/reporte_pagos.html', context)


def exportar_reporte_excel(pagos, totales, año):
    """Genera archivo Excel con el reporte de pagos."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos Airbnb"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.merge_cells('A1:J1')
    ws['A1'] = 'REPORTE DE INGRESOS AIRBNB - PLATAFORMAS TECNOLÓGICAS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = f'Año: {año} | Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = [
        'Código', 'Huésped', 'Anuncio', 'Check-in', 'Check-out',
        'Bruto', 'Comisión', 'ISR 4%', 'IVA 8%', 'Neto'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    row_num = 5
    for pago in pagos:
        ws.cell(row=row_num, column=1, value=pago.codigo_confirmacion or '-')
        ws.cell(row=row_num, column=2, value=pago.huesped)
        ws.cell(row=row_num, column=3, value=pago.anuncio.nombre if pago.anuncio else '-')
        ws.cell(row=row_num, column=4, value=pago.fecha_checkin.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=5, value=pago.fecha_checkout.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=6, value=float(pago.monto_bruto))
        ws.cell(row=row_num, column=7, value=float(pago.comision_airbnb))
        ws.cell(row=row_num, column=8, value=float(pago.retencion_isr))
        ws.cell(row=row_num, column=9, value=float(pago.retencion_iva))
        ws.cell(row=row_num, column=10, value=float(pago.monto_neto))

        for col in range(1, 11):
            ws.cell(row=row_num, column=col).border = border

        row_num += 1

    total_row = row_num
    ws.cell(row=total_row, column=5, value="TOTALES:").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=float(totales['total_bruto'] or 0)).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=float(totales['total_comision'] or 0)).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=float(totales['total_isr'] or 0)).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=float(totales['total_iva'] or 0)).font = Font(bold=True)
    ws.cell(row=total_row, column=10, value=float(totales['total_neto'] or 0)).font = Font(bold=True)

    nota_row = total_row + 2
    ws.merge_cells(f'A{nota_row}:J{nota_row}')
    ws[f'A{nota_row}'] = 'Régimen: Actividad Empresarial - Plataformas Tecnológicas (Art. 113-A LISR)'
    ws[f'A{nota_row}'].font = Font(italic=True, color="666666")

    column_widths = [15, 25, 20, 12, 12, 12, 12, 10, 10, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Airbnb_Pagos_{año}.xlsx"'
    wb.save(response)

    return response


# ==========================================
# ICAL INVERSO - Exportar eventos del ERP
# ==========================================
@rate_limit(key='ical_eventos', limit=120, window=60)
def generar_ical_eventos(request):
    """
    Genera un archivo iCal con los eventos confirmados del ERP.
    Esta URL se importa en Airbnb para bloquear fechas automáticamente.

    URL: /airbnb/ical/eventos/?token=XXX

    Exige ICAL_PUBLIC_TOKEN siempre (fail-closed). Antes, si la variable no
    estaba configurada la validación entera se saltaba y el feed quedaba
    abierto a internet publicando el nombre de cada cliente con su evento y
    su fecha — que fue exactamente lo que pasó en producción. Sin token
    configurado la respuesta es 403: es preferible que Airbnb deje de
    sincronizar, algo visible y reparable, a exponer la cartera en silencio.
    """
    import hmac

    from django.conf import settings
    from django.http import HttpResponseForbidden

    from comercial.models import Cotizacion

    token_esperado = getattr(settings, 'ICAL_PUBLIC_TOKEN', '')
    if not token_esperado:
        logger.error(
            'ICAL_PUBLIC_TOKEN no está configurado: se rechaza toda petición al feed iCal.'
        )
        return HttpResponseForbidden('Feed no disponible')

    if not hmac.compare_digest(request.GET.get('token', ''), token_esperado):
        return HttpResponseForbidden('Token inválido')

    # Solo eventos confirmados (últimos 30 días + futuros)
    cotizaciones = Cotizacion.objects.filter(
        estado='CONFIRMADA',
        fecha_evento__gte=timezone.now().date() - timedelta(days=30)
    ).select_related('cliente')

    # Generar contenido iCal
    lineas = [
        'BEGIN:VCALENDAR',
        'VERSION:2.0',
        'PRODID:-//Quinta Koox Tanil//ERP//ES',
        'CALSCALE:GREGORIAN',
        'METHOD:PUBLISH',
        'X-WR-CALNAME:Eventos Quinta Koox Tanil',
        'X-WR-TIMEZONE:America/Merida',
    ]

    for cot in cotizaciones:
        uid = f"evento-{cot.id}@qkt-erp"

        # Fecha de inicio y fin (evento de día completo)
        fecha_inicio = cot.fecha_evento.strftime('%Y%m%d')
        fecha_fin = (cot.fecha_evento + timedelta(days=1)).strftime('%Y%m%d')

        # Timestamp de creación
        dtstamp = timezone.now().strftime('%Y%m%dT%H%M%SZ')
        created = cot.created_at.strftime('%Y%m%dT%H%M%SZ') if hasattr(cot, 'created_at') and cot.created_at else dtstamp

        # Ni el nombre del cliente ni el del evento salen del ERP: el feed solo
        # existe para que Airbnb bloquee la fecha, y para eso basta el folio.
        # Publicar "Cliente: NOMBRE" regalaba la cartera a quien tuviera la URL.
        # Como el folio es numérico, tampoco hay nada que escapar (RFC 5545).
        titulo = f"EVENTO QKT: COT-{cot.id:03d}"
        descripcion = "Fecha no disponible."

        lineas.extend([
            'BEGIN:VEVENT',
            f'UID:{uid}',
            f'DTSTAMP:{dtstamp}',
            f'CREATED:{created}',
            f'DTSTART;VALUE=DATE:{fecha_inicio}',
            f'DTEND;VALUE=DATE:{fecha_fin}',
            f'SUMMARY:{titulo}',
            f'DESCRIPTION:{descripcion}',
            'STATUS:CONFIRMED',
            'TRANSP:OPAQUE',
            'END:VEVENT',
        ])

    lineas.append('END:VCALENDAR')

    contenido = '\r\n'.join(lineas)

    response = HttpResponse(contenido, content_type='text/calendar; charset=utf-8')
    response['Content-Disposition'] = 'inline; filename="eventos_qkt.ics"'

    return response


# ==========================================
# BLOQUEO MANUAL EN AIRBNB
# ==========================================
@staff_member_required
@permission_required('airbnb.view_anuncioairbnb', raise_exception=True)
def bloquear_en_airbnb(request, cotizacion_id):
    """
    Redirige a Airbnb para bloquear manualmente las fechas de un evento.
    Abre el calendario del anuncio en la fecha específica.
    """
    from comercial.models import Cotizacion

    cotizacion = get_object_or_404(Cotizacion, pk=cotizacion_id)

    # Obtener anuncios que afectan la quinta
    anuncios = AnuncioAirbnb.objects.filter(
        afecta_eventos_quinta=True,
        activo=True
    )

    if not anuncios.exists():
        messages.warning(request, " No hay anuncios configurados que afecten la quinta")
        return redirect('admin:comercial_cotizacion_change', cotizacion_id)

    # Generar URLs de bloqueo para cada anuncio
    urls_bloqueo = []

    for anuncio in anuncios:
        if anuncio.airbnb_listing_id:
            # URL directa al multicalendar de Airbnb México
            url = f"https://www.airbnb.mx/multicalendar/{anuncio.airbnb_listing_id}"
            urls_bloqueo.append({
                'nombre': anuncio.nombre,
                'url': url,
                'listing_id': anuncio.airbnb_listing_id,
            })

    if not urls_bloqueo:
        messages.warning(request, " Los anuncios no tienen Listing ID configurado")
        return redirect('admin:comercial_cotizacion_change', cotizacion_id)

    if len(urls_bloqueo) == 1:
        # Si solo hay un anuncio, redirigir directamente
        messages.info(
            request,
            f" Bloquea la fecha {cotizacion.fecha_evento.strftime('%d/%m/%Y')} en el calendario de Airbnb"
        )
        return redirect(urls_bloqueo[0]['url'])

    # Si hay múltiples anuncios, mostrar página con links
    context = {
        'cotizacion': cotizacion,
        'urls_bloqueo': urls_bloqueo,
        'title': f'Bloquear en Airbnb: {cotizacion.nombre_evento}',
    }

    return render(request, 'admin/airbnb/bloquear_manual.html', context)


@staff_member_required
@permission_required('airbnb.view_anuncioairbnb', raise_exception=True)
def dashboard_airbnb(request):
    """
    Dashboard del módulo Airbnb con estadísticas y accesos rápidos.
    """
    hoy = timezone.now()

    # Estadísticas
    total_anuncios = AnuncioAirbnb.objects.filter(activo=True).count()

    pagos_mes = PagoAirbnb.objects.filter(
        fecha_checkin__year=hoy.year,
        fecha_checkin__month=hoy.month,
        estado='PAGADO'
    ).aggregate(
        total_neto=Sum('monto_neto'),
        num_reservas=Count('id'),
    )

    conflictos_count = ConflictoCalendario.objects.filter(estado='PENDIENTE').count()

    # Próximas reservas
    reservas_proximas = ReservaAirbnb.objects.filter(
        fecha_inicio__gte=hoy.date(),
        estado='CONFIRMADA',
        anuncio__activo=True
    ).select_related('anuncio').order_by('fecha_inicio')[:5]

    # Conflictos pendientes
    conflictos = ConflictoCalendario.objects.filter(
        estado='PENDIENTE'
    ).select_related('reserva_airbnb__anuncio', 'cotizacion')[:5]

    # Datos para gráfica de ingresos (últimos 6 meses)
    ingresos_labels = []
    ingresos_data = []

    for i in range(5, -1, -1):
        if i > 0:
            fecha = hoy - timedelta(days=30*i)
        else:
            fecha = hoy

        total = PagoAirbnb.objects.filter(
            fecha_checkin__year=fecha.year,
            fecha_checkin__month=fecha.month,
            estado='PAGADO'
        ).aggregate(total=Sum('monto_neto'))['total'] or 0

        ingresos_labels.append(fecha.strftime('%b %Y'))
        ingresos_data.append(float(total))

    # URL del iCal
    ical_url = request.build_absolute_uri('/airbnb/ical/eventos/')

    context = {
        'total_anuncios': total_anuncios,
        'pagos_mes': pagos_mes,
        'conflictos_count': conflictos_count,
        'reservas_proximas': reservas_proximas,
        'conflictos': conflictos,
        'today': hoy.date(),
        # Sin json.dumps: los serializa la plantilla con |json_script.
        'ingresos_labels': ingresos_labels,
        'ingresos_data': ingresos_data,
        'ical_url': ical_url,
        'title': 'Dashboard Airbnb',
    }

    return render(request, 'admin/airbnb/dashboard.html', context)


@staff_member_required
@permission_required('airbnb.view_depositoconciliado', raise_exception=True)
def conciliacion_depositos_airbnb(request):
    """
    Cuadra los depósitos de Airbnb del mes contra los abonos del banco.

    Airbnb junta en un solo payout las reservas que liquida el mismo día, así
    que la conciliación es por depósito y no por reserva: el estado de cuenta
    trae un abono por payout, no uno por huésped. Antes esto se cuadraba a
    mano contra el PDF del banco.

    El POST resuelve los depósitos ambiguos: cuando el banco no conservó el id
    del payout y dos abonos encajan igual de bien, el sistema no elige —quien
    concilia dice cuál es cuál y la decisión queda guardada—.
    """
    from .services import ConciliacionDepositosService

    hoy = timezone.localdate()

    def _entero(nombre, defecto, minimo, maximo):
        crudo = str(request.GET.get(nombre, defecto)).replace(',', '').strip()
        try:
            valor = int(crudo)
        except (TypeError, ValueError):
            return defecto
        return valor if minimo <= valor <= maximo else defecto

    mes = _entero('mes', hoy.month, 1, 12)
    anio = _entero('anio', hoy.year, 2000, 2100)

    if request.method == 'POST':
        return _resolver_deposito_airbnb(request, mes, anio)

    servicio = ConciliacionDepositosService(mes=mes, anio=anio)
    depositos = servicio.conciliar()

    context = {
        'title': 'Conciliación de depósitos Airbnb',
        'depositos': depositos,
        'totales': servicio.totales(depositos),
        'mes': mes,
        'anio': anio,
        'mes_nombre': MESES[mes],
        'meses': sorted(MESES.items()),
        'anios': list(range(hoy.year - 2, hoy.year + 1)),
    }
    return render(request, 'admin/airbnb/conciliacion_depositos.html', context)


def _resolver_deposito_airbnb(request, mes, anio):
    """Confirma o suelta el abono de un payout, y vuelve al mismo mes."""
    from contabilidad.models import MovimientoEstadoCuenta

    from .services import ConciliacionDepositosService

    destino = (f"{reverse('conciliacion_depositos_airbnb')}"
               f"?mes={mes}&anio={anio}")
    payout_id = (request.POST.get('payout_id') or '').strip()
    if not payout_id:
        messages.error(request, "Falta el depósito a conciliar.")
        return redirect(destino)

    if request.POST.get('deshacer'):
        if ConciliacionDepositosService.deshacer(payout_id):
            messages.success(
                request, f"Se soltó el emparejamiento del depósito {payout_id}.")
        return redirect(destino)

    try:
        movimiento = MovimientoEstadoCuenta.objects.get(
            pk=int(request.POST.get('movimiento_id', '')), abono__gt=0)
    except (ValueError, TypeError, MovimientoEstadoCuenta.DoesNotExist):
        messages.error(request, "El movimiento bancario seleccionado no existe.")
        return redirect(destino)

    ConciliacionDepositosService.confirmar(payout_id, movimiento, request.user)
    messages.success(
        request,
        f"Depósito {payout_id} conciliado con el abono del "
        f"{movimiento.fecha:%d/%m/%Y} por ${movimiento.abono}.")
    return redirect(destino)


@staff_member_required
@permission_required('airbnb.view_pagoairbnb', raise_exception=True)
def reporte_fiscal_airbnb(request):
    """
    Reporte fiscal mensual de ingresos Airbnb.

    El período se determina por FECHA DE PAGO, no por check-in: para el
    régimen de plataformas tecnológicas lo que fija el mes es cuándo la
    plataforma pagó y retuvo. Agrupar por check-in metía en agosto una reserva
    que Airbnb liquidó en septiembre, y el reporte no cuadraba con la
    declaración ni con la constancia de retenciones.
    """
    from django.template.loader import render_to_string
    from weasyprint import HTML

    from .models import AnuncioAirbnb, PagoAirbnb

    # `localdate()` y no `now().date()`: con TIME_ZONE='America/Merida' el mes
    # cambiaba seis horas antes de tiempo (mismo bug ya corregido en comercial).
    hoy = timezone.localdate()

    # Parámetros defensivos: antes un `?mes=abc` o `?mes=13` reventaba con un
    # ValueError o un KeyError y devolvía un 500 desde la URL.
    def _entero(nombre, defecto, minimo, maximo):
        crudo = str(request.GET.get(nombre, defecto)).replace(',', '').strip()
        try:
            valor = int(crudo)
        except (TypeError, ValueError):
            return defecto
        return valor if minimo <= valor <= maximo else defecto

    mes = _entero('mes', hoy.month, 1, 12)
    anio = _entero('anio', hoy.year, 2000, 2100)

    pagos = (PagoAirbnb.objects
             .filter(fecha_pago__month=mes, fecha_pago__year=anio)
             .exclude(estado__in=('CANCELADO', 'REEMBOLSADO'))
             .select_related('anuncio')
             .order_by('anuncio', 'fecha_pago'))

    agregado = pagos.aggregate(
        bruto=Sum('monto_bruto'), comision=Sum('comision_airbnb'),
        isr=Sum('retencion_isr'), iva=Sum('retencion_iva'),
        ish=Sum('impuesto_hospedaje'), iva_trasladado=Sum('iva_trasladado'),
        neto=Sum('monto_neto'),
    )

    def _d(valor):
        return valor or Decimal('0.00')

    bruto = _d(agregado['bruto'])
    lista = list(pagos)
    noches_vendidas = sum(p.noches for p in lista)

    totales = {
        'bruto': bruto,
        'comision': _d(agregado['comision']),
        'isr': _d(agregado['isr']),
        'iva': _d(agregado['iva']),
        'ish': _d(agregado['ish']),
        'iva_trasladado': _d(agregado['iva_trasladado']),
        'neto': _d(agregado['neto']),
        'num_reservas': len(lista),
        'noches': noches_vendidas,
        'ingreso_actividad': bruto - _d(agregado['comision']),
        'total_retenciones': _d(agregado['isr']) + _d(agregado['iva']),
        'tarifa_promedio': _promedio(bruto, noches_vendidas),
    }

    # Los pagos cuyo neto no cuadra se listan aparte: declarar sobre un
    # registro descuadrado es justo lo que no debe pasar inadvertido.
    totales['descuadrados'] = [p for p in lista if not p.cuadra]

    dias_del_mes = calendar.monthrange(anio, mes)[1]
    anuncios = list(AnuncioAirbnb.objects.filter(activo=True))

    # Un solo recorrido en memoria en vez de dos consultas por anuncio.
    por_anuncio = defaultdict(list)
    for pago in lista:
        por_anuncio[pago.anuncio_id].append(pago)

    resumen_propiedades = []
    for anuncio in anuncios:
        propios = por_anuncio.get(anuncio.id)
        if not propios:
            continue
        bruto_prop = sum((p.monto_bruto for p in propios), Decimal('0.00'))
        noches_prop = sum(p.noches for p in propios)
        resumen_propiedades.append({
            'nombre': anuncio.nombre,
            'num_reservas': len(propios),
            'noches': noches_prop,
            'tarifa_promedio': _promedio(bruto_prop, noches_prop),
            'bruto': bruto_prop,
            'comision': sum((p.comision_airbnb for p in propios), Decimal('0.00')),
            'isr': sum((p.retencion_isr for p in propios), Decimal('0.00')),
            'iva': sum((p.retencion_iva for p in propios), Decimal('0.00')),
            'neto': sum((p.monto_neto for p in propios), Decimal('0.00')),
            'porcentaje': _porcentaje(bruto_prop, bruto),
            # Métricas estándar de hospedaje. La ocupación necesita las noches
            # DISPONIBLES (días del mes por anuncio), que es lo que faltaba
            # para poder interpretar la tarifa promedio.
            'ocupacion': _porcentaje(Decimal(noches_prop), Decimal(dias_del_mes)),
            'adr': _promedio(bruto_prop, noches_prop),
            'revpar': _promedio(bruto_prop, dias_del_mes),
        })

    noches_disponibles = dias_del_mes * len(anuncios)
    totales['ocupacion'] = _porcentaje(Decimal(noches_vendidas),
                                       Decimal(noches_disponibles))
    totales['adr'] = _promedio(bruto, noches_vendidas)
    totales['revpar'] = _promedio(bruto, noches_disponibles)

    context = {
        'pagos': lista,
        'totales': totales,
        'resumen_propiedades': resumen_propiedades,
        'mes': mes,
        'anio': anio,
        'mes_nombre': MESES[mes],
        'total_propiedades': len(anuncios),
        'dias_del_mes': dias_del_mes,
        'fecha_generacion': hoy.strftime('%d/%m/%Y'),
    }

    html_string = render_to_string('airbnb/reporte_fiscal_airbnb.html', context)
    pdf_bytes = HTML(string=html_string).write_pdf()

    filename = f"Reporte_Fiscal_Airbnb_{MESES[mes]}_{anio}.pdf"
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


def _promedio(total: Decimal, divisor) -> Decimal:
    if not divisor:
        return Decimal('0.00')
    return (Decimal(total) / Decimal(divisor)).quantize(Decimal('0.01'))


def _porcentaje(parte: Decimal, total: Decimal) -> Decimal:
    if not total:
        return Decimal('0.0')
    return (Decimal(parte) / Decimal(total) * 100).quantize(Decimal('0.1'))
